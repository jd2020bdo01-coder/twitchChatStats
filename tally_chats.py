import re
import pandas as pd
import os
from datetime import datetime, timedelta, date
from collections import defaultdict
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
import sys

def tally_chats(log_path):
    chat_counts = {}
    with open(log_path, 'r', encoding='utf-8') as f:
        for line in f:
            # Extract username after timestamp, e.g. [15:28:55] username: message
            match = re.match(r'^\[\d{2}:\d{2}:\d{2}\] ([^:]+):', line)
            if match:
                name = match.group(1).strip()
                chat_counts[name] = chat_counts.get(name, 0) + 1
    return chat_counts

def collect_user_times_from_counts(channel_path):
    user_times = defaultdict(list)
    for log_file in os.listdir(channel_path):
        if log_file.endswith('.log'):
            with open(os.path.join(channel_path, log_file), encoding='utf-8') as f:
                for line in f:
                    match = re.match(r'^\[(\d{2}:\d{2}:\d{2})\] ([^:]+):', line)
                    if match:
                        time_str = match.group(1)
                        user = match.group(2).strip()
                        date_match = re.search(r'(\d{4}-\d{2}-\d{2})', log_file)
                        if date_match:
                            dt_str = f"{date_match.group(1)} {time_str}"
                            dt = datetime.strptime(dt_str, "%Y-%m-%d %H:%M:%S")
                            user_times[user].append(dt)
    return user_times

def has_overlap(times1, times2, threshold=timedelta(seconds=2)):
    i, j = 0, 0
    times1 = sorted(times1)
    times2 = sorted(times2)
    while i < len(times1) and j < len(times2):
        if abs((times1[i] - times2[j]).total_seconds()) <= threshold.total_seconds():
            return True
        if times1[i] < times2[j]:
            i += 1
        else:
            j += 1
    return False

def group_unique_users(user_times):
    users = list(user_times.keys())
    n = len(users)
    groups = []
    assigned = set()
    for i in range(n):
        if users[i] in assigned:
            continue
        group = [users[i]]
        for j in range(i+1, n):
            if users[j] not in assigned and not has_overlap(user_times[users[i]], user_times[users[j]]):
                group.append(users[j])
                assigned.add(users[j])
        assigned.add(users[i])
        groups.append(group)
    return groups

def collect_user_messages(channel_path, date_filter=None):
    user_messages = defaultdict(list)
    for log_file in os.listdir(channel_path):
        if log_file.endswith('.log'):
            # Date filter logic
            if date_filter:
                date_match = re.search(r'(\d{4}-\d{2}-\d{2})', log_file)
                if date_match:
                    log_date = date_match.group(1)
                    if ':' in date_filter:
                        start, end = date_filter.split(':')
                        if not (start <= log_date <= end):
                            continue
                    else:
                        if log_date != date_filter:
                            continue
                else:
                    continue
            with open(os.path.join(channel_path, log_file), encoding='utf-8') as f:
                for line in f:
                    match = re.match(r'^\[\d{2}:\d{2}:\d{2}\] ([^:]+): (.*)$', line)
                    if match:
                        user = match.group(1).strip()
                        message = match.group(2).strip()
                        user_messages[user].append(message)
    return user_messages

def group_users_by_stylometry(user_messages, similarity_threshold=0.6):
    users = list(user_messages.keys())
    if len(users) <= 1:
        return [[u] for u in users], {u: 0.0 for u in users}, {u: [] for u in users}
    corpus = [' '.join(user_messages[u]) for u in users]
    vectorizer = TfidfVectorizer()
    X = vectorizer.fit_transform(corpus)
    sim_matrix = cosine_similarity(X)
    n = len(users)
    groups = []
    assigned = set()
    alt_scores = {u: 0.0 for u in users}
    similar_users = {u: [] for u in users}
    for i in range(n):
        if users[i] in assigned:
            continue
        group = [users[i]]
        for j in range(n):
            if i != j:
                sim = sim_matrix[i, j]
                if sim >= similarity_threshold:
                    similar_users[users[i]].append(f"{users[j]} ({round(sim*100,1)}%)")
                alt_scores[users[i]] = max(alt_scores[users[i]], sim)
        # Grouping logic (unchanged)
        for j in range(i+1, n):
            if users[j] not in assigned and sim_matrix[i, j] >= similarity_threshold:
                group.append(users[j])
                assigned.add(users[j])
        assigned.add(users[i])
        groups.append(group)
    # Normalize scores to percentage
    for u in alt_scores:
        alt_scores[u] = round(alt_scores[u] * 100, 1)
    return groups, alt_scores, similar_users

def save_to_excel_with_dates(chat_counts, output_path, start_date, end_date, channel, unique_user_count, alt_likelihoods=None, similar_users=None):
    import openpyxl
    from openpyxl.styles import Font
    # Prepare data
    date_range_row = pd.DataFrame({
        'User': ['Date Range:'],
        'Chat Count': [f'{start_date} to {end_date}'],
        'Alt Likelihood': [''],
        'Similar Users': ['']
    })
    total_unique_users_row = pd.DataFrame({
        'User': ['Possible unique users:'],
        'Chat Count': [unique_user_count],
        'Alt Likelihood': [''],
        'Similar Users': ['']
    })
    headers = pd.DataFrame({
        'User': ['User'],
        'Chat Count': ['Chat Count'],
        'Alt Likelihood': ['Alt Likelihood'],
        'Similar Users': ['Similar Users']
    })
    # Prepare user data with alt likelihoods and similar users
    data = []
    for user, count in chat_counts.items():
        likelihood = ''
        if alt_likelihoods and user in alt_likelihoods:
            likelihood = alt_likelihoods[user]
        similar = ''
        if similar_users and user in similar_users:
            similar = ', '.join(similar_users[user])
        data.append({'User': user, 'Chat Count': count, 'Alt Likelihood': likelihood, 'Similar Users': similar})
    df = pd.DataFrame(data)
    df = df.sort_values(by='Chat Count', ascending=False)
    # Concatenate: date range, total unique users, headers, then data
    df = pd.concat([date_range_row, total_unique_users_row, headers, df], ignore_index=True)
    df.to_excel(output_path, index=False, header=False)

    # Highlight the date range row, total unique users row, and headers
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    bold_font = Font(bold=True)
    for cell in ws[1]:  # Date range row
        cell.font = bold_font
    for cell in ws[2]:  # Total unique users row
        cell.font = bold_font
    for cell in ws[3]:  # Header row
        cell.font = bold_font
    wb.save(output_path)

def tally_chats_in_folder(channels_dir, output_root, channel_filter=None, date_filter=None):
    for channel in os.listdir(channels_dir):
        if channel_filter and channel != channel_filter:
            continue
        channel_path = os.path.join(channels_dir, channel)
        if os.path.isdir(channel_path):
            chat_counts = {}
            log_dates = []
            for log_file in os.listdir(channel_path):
                if log_file.endswith('.log'):
                    # Extract date from filename, expects format: channel-YYYY-MM-DD.log
                    date_match = re.search(r'(\d{4}-\d{2}-\d{2})', log_file)
                    if date_match:
                        log_date = date_match.group(1)
                        # Date filter logic
                        if date_filter:
                            if ':' in date_filter:
                                start, end = date_filter.split(':')
                                if not (start <= log_date <= end):
                                    continue
                            else:
                                if log_date != date_filter:
                                    continue
                        log_dates.append(log_date)
                    else:
                        if date_filter:
                            continue
                    log_path = os.path.join(channel_path, log_file)
                    file_counts = tally_chats(log_path)
                    for user, count in file_counts.items():
                        chat_counts[user] = chat_counts.get(user, 0) + count
            if chat_counts:
                # If date_filter is set, use it for the date range row
                if date_filter:
                    if ':' in date_filter:
                        start_date, end_date = date_filter.split(':')
                    else:
                        start_date = end_date = date_filter
                elif log_dates:
                    start_date = min(log_dates)
                    end_date = max(log_dates)
                else:
                    start_date = end_date = 'Unknown'
                # Use stylometry-based unique user grouping
                user_messages = collect_user_messages(channel_path, date_filter)
                stylometry_groups, alt_scores, similar_users = group_users_by_stylometry(user_messages)
                unique_user_count = len(stylometry_groups)
                # Compute alt likelihoods: use percentage
                alt_likelihoods = {user: f"{alt_scores[user]}%" for user in chat_counts}
                # Build output filename with date filter
                filename = f"{channel}_tally"
                if date_filter:
                    if ':' in date_filter:
                        filename += f"_{date_filter.replace(':','-')}"
                    else:
                        filename += f"_{date_filter}"
                output_file = os.path.join(output_root, f"{filename}.xlsx")
                save_to_excel_with_dates(chat_counts, output_file, start_date, end_date, channel, unique_user_count, alt_likelihoods, similar_users)
                print(f"Tally for {channel} saved to {output_file}")

if __name__ == "__main__":
    channels_dir = "Channels"
    output_root = "."
    channel_filter = None
    date_filter = None
    if len(sys.argv) > 1:
        channel_filter = sys.argv[1]
    if len(sys.argv) > 2:
        date_filter = sys.argv[2]
    tally_chats_in_folder(channels_dir, output_root, channel_filter, date_filter)
